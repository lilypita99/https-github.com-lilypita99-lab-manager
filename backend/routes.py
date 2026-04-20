import re
import os
import json
import base64
import secrets as _sec
import smtplib
import urllib.parse
from datetime import datetime
from functools import wraps
from email.message import EmailMessage

import requests as _http
from flask import jsonify, request, redirect
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from backend.models import AuthToken, User, LabGroup, LabInvite, db, InventoryItem, PurchaseHistory

# In-memory OAuth state store for CSRF protection (token → provider)
_oauth_states: dict = {}


def _get_current_user():
    """Return User from Bearer token in Authorization header, or None."""
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        return None
    token_str = auth_header[7:]
    token = AuthToken.query.filter_by(token=token_str).first()
    if token is None:
        return None
    return token.user


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = _get_current_user()
        if user is None:
            return jsonify({"error": "Authentication required"}), 401
        return f(*args, **kwargs)
    return decorated


def _send_invite_email(inviter, group, invite_email, invite_role):
    smtp_host = os.environ.get("LABFLOW_SMTP_HOST", "").strip()
    if not smtp_host:
        return False, "Email delivery not configured. Set LABFLOW_SMTP_HOST to enable sending."

    smtp_port = int(os.environ.get("LABFLOW_SMTP_PORT", "587"))
    smtp_user = os.environ.get("LABFLOW_SMTP_USER", "").strip()
    smtp_pass = os.environ.get("LABFLOW_SMTP_PASS", "").strip()
    smtp_from = os.environ.get("LABFLOW_SMTP_FROM", "LabFlow <no-reply@labflow.local>").strip()
    smtp_use_tls = os.environ.get("LABFLOW_SMTP_USE_TLS", "true").strip().lower() in {
        "1", "true", "yes", "y"
    }

    inviter_name = inviter.display_name or inviter.email
    role_label = "Manager" if invite_role == "manager" else "Lab User"

    message = EmailMessage()
    message["Subject"] = f"You're invited to join {group.group_name} on LabFlow"
    message["From"] = smtp_from
    message["To"] = invite_email
    message.set_content(
        f"Hello,\n\n"
        f"{inviter_name} invited you to join '{group.group_name}' in '{group.lab_name}' on LabFlow "
        f"as a {role_label}.\n\n"
        f"To join, register or log in with this email address ({invite_email}) in the LabFlow app, "
        f"then review and accept the invite in the app.\n\n"
        f"Thanks,\nLabFlow"
    )

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            if smtp_use_tls:
                server.starttls()
            if smtp_user:
                server.login(smtp_user, smtp_pass)
            server.send_message(message)
        return True, "Invite email sent."
    except Exception as exc:
        return False, f"Invite saved, but email failed to send: {exc}"


def normalize_header(value):
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(value).strip().lower())


def to_float(value, default=0.0):
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def to_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "need", "needed"}


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def recalculate_need_fields(item, preserve_in_process=False):
    actual = float(item.actual_quantity or 0)
    desired = float(item.desired_quantity or 0)
    need = max(0.0, desired - actual)
    item.order_quantity = need
    item.need_to_order = need > 0
    if preserve_in_process and item.order_status == "in_process":
        return
    item.order_status = "needs" if need > 0 else "in_stock"


def create_purchase_history_entry(item, quantity, ordered_at):
    entry = PurchaseHistory(
        inventory_item_id=item.id,
        item_name=item.name,
        category=item.category,
        catalog_number=item.catalog_number,
        vendor=item.vendor,
        requested_by=item.requested_by,
        lab_name=item.lab_name,
        quantity_ordered=quantity,
        quantity_received=0,
        status="ordered",
        ordered_at=ordered_at,
    )
    db.session.add(entry)
    return entry


def find_open_purchase_history_entry(item_id):
    return PurchaseHistory.query.filter(
        PurchaseHistory.inventory_item_id == item_id,
        PurchaseHistory.received_at.is_(None),
    ).order_by(PurchaseHistory.ordered_at.desc(), PurchaseHistory.id.desc()).first()


HEADER_ALIASES = {
    "name": {"item", "itemname", "name", "product", "itemdescription"},
    "category": {"category", "cat", "type", "itemtype", "producttype", "section", "group"},
    "catalog_number": {"catalog", "catalognumber", "catalog#", "catno", "catalogno", "catalognum", "catalogn"},
    "location_stored": {"location", "locationstored", "storedat", "storage", "room"},
    "actual_quantity": {"actualquantity", "quantity", "qty", "currentquantity", "onhand"},
    "desired_quantity": {"desiredquantity", "suggestedquantity", "targetquantity", "par", "desiredqty"},
    "need_to_order": {"needtoorder", "orderneeded", "needorder", "needsorder"},
    "order_quantity": {"orderquantity", "qtytoorder", "ordertqty", "reorderquantity", "orderqty"},
}

PURCHASE_HISTORY_HEADER_ALIASES = {
    "item_name": {"item", "itemname", "name", "product", "itemdescription"},
    "category": {"category", "cat", "type", "itemtype", "producttype", "section", "group"},
    "catalog_number": {"catalog", "catalognumber", "catalog#", "catno", "catalogno", "catalognum", "catalogn"},
    "vendor": {"vendor", "supplier", "company"},
    "requested_by": {"requestedby", "requestor", "requester", "requestedbyname"},
    "lab_name": {"lab", "labname", "project", "projectname", "groupname"},
    "quantity_ordered": {"quantityordered", "orderedqty", "orderedquantity", "orderqty", "qtyordered"},
    "quantity_received": {"quantityreceived", "receivedqty", "receivedquantity", "qtyreceived"},
    "status": {"status", "orderstatus"},
    "ordered_at": {"orderedat", "orderedon", "orderdate", "dateordered", "timestamp", "datetime", "date", "submittedat", "requesttimestamp"},
    "received_at": {"receivedat", "receivedon", "receiveddate", "datereceived"},
}


def find_column_indexes(header_row):
    indexed = {}
    for i, value in enumerate(header_row):
        key = normalize_header(value)
        for field, aliases in HEADER_ALIASES.items():
            if key in aliases and field not in indexed:
                indexed[field] = i
    return indexed


def find_column_indexes_for_aliases(header_row, aliases_map):
    indexed = {}
    for i, value in enumerate(header_row):
        key = normalize_header(value)
        for field, aliases in aliases_map.items():
            if key in aliases and field not in indexed:
                indexed[field] = i
    return indexed


def parse_datetime_cell(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value)
        except Exception:
            return None
    text = str(value).strip()
    if not text:
        return None

    # Try ISO-style timestamps first (e.g. 2026-05-06T00:00:00 or 2026-05-06 00:00:00).
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        pass

    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d",
        "%m-%d-%Y %H:%M:%S",
        "%m-%d-%Y %H:%M",
        "%m-%d-%Y",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
        "%m/%d/%y %H:%M:%S",
        "%m/%d/%y %H:%M",
        "%m/%d/%y",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def is_likely_history_data_row(row):
    if row is None:
        return False
    if all(cell is None or str(cell).strip() == "" for cell in row):
        return False

    first = row[0] if len(row) > 0 else None
    # Most user sheets place date in the first column.
    if parse_datetime_cell(first) is None:
        return False

    # Typical layouts have a description/item text around column F (index 5).
    maybe_item = row[5] if len(row) > 5 else None
    return maybe_item is not None and str(maybe_item).strip() != ""


def get_row_value(row, indexes, field, default=None):
    if field not in indexes:
        return default
    idx = indexes[field]
    if idx >= len(row):
        return default
    return row[idx]


def make_item_key(name, catalog_number, location_stored):
    return (
        str(name or "").strip().lower(),
        str(catalog_number or "").strip().lower(),
        str(location_stored or "").strip().lower(),
    )


def register_routes(app):

    # ── Auth routes ────────────────────────────────────────────────────────────

    # Email/password auth is disabled — sign-in is OAuth-only (Google / Microsoft).
    @app.route("/api/auth/register", methods=["POST"])
    def auth_register():
        return jsonify({"error": "Registration via email is disabled. Please sign in with Google or Outlook."}), 404

    @app.route("/api/auth/login", methods=["POST"])
    def auth_login():
        return jsonify({"error": "Password login is disabled. Please sign in with Google or Outlook."}), 404

    @app.route("/api/auth/me", methods=["GET"])
    def auth_me():
        user = _get_current_user()
        if user is None:
            return jsonify({"error": "Not authenticated"}), 401
        return jsonify({"user": user.to_dict()})

    @app.route("/api/auth/logout", methods=["POST"])
    def auth_logout():
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token_str = auth_header[7:]
            token = AuthToken.query.filter_by(token=token_str).first()
            if token:
                db.session.delete(token)
                db.session.commit()
        return jsonify({"success": True})

    # ── OAuth routes ───────────────────────────────────────────────────────────

    def _google_client_config_error():
        client_id = os.environ.get("GOOGLE_CLIENT_ID", "").strip()
        client_secret = os.environ.get("GOOGLE_CLIENT_SECRET", "").strip()
        if (
            not client_id
            or client_id.startswith("your_")
            or not client_id.endswith(".apps.googleusercontent.com")
        ):
            return "Google OAuth is misconfigured: GOOGLE_CLIENT_ID is missing or invalid.", None, None
        if not client_secret or client_secret.startswith("your_"):
            return "Google OAuth is misconfigured: GOOGLE_CLIENT_SECRET is missing or invalid.", None, None
        return None, client_id, client_secret

    def _oauth_finish(email, name, provider, frontend_url):
        """Find or create user from OAuth, issue token, redirect to frontend."""
        if not email:
            return redirect(f"{frontend_url}?oauth_error=no_email")
        user = User.query.filter_by(email=email).first()
        if user is None:
            user = User(email=email, display_name=name or email, role="lab")
            user.set_password(_sec.token_urlsafe(32))
            db.session.add(user)
            db.session.flush()
        token_str = AuthToken.generate(user.id)
        user_b64 = base64.urlsafe_b64encode(json.dumps(user.to_dict()).encode()).decode()
        return redirect(f"{frontend_url}?oauth_token={urllib.parse.quote(token_str)}&oauth_user={user_b64}")

    @app.route("/api/auth/google/client", methods=["GET"])
    def google_client_config():
        config_error, client_id, _ = _google_client_config_error()
        if config_error:
            return jsonify({"error": config_error}), 503
        return jsonify({"client_id": client_id})

    @app.route("/api/auth/google/id-token", methods=["POST"])
    def google_id_token_login():
        config_error, client_id, _ = _google_client_config_error()
        if config_error:
            return jsonify({"error": config_error}), 503

        payload = request.get_json(silent=True) or {}
        id_token = str(payload.get("id_token", "")).strip()
        if not id_token:
            return jsonify({"error": "Missing id_token"}), 400

        try:
            verify_res = _http.get(
                "https://oauth2.googleapis.com/tokeninfo",
                params={"id_token": id_token},
                timeout=10,
            )
            if not verify_res.ok:
                return jsonify({"error": "Invalid Google token"}), 401

            token_info = verify_res.json()
            aud = token_info.get("aud", "")
            if aud != client_id:
                return jsonify({"error": "Token audience mismatch"}), 401

            if token_info.get("email_verified") not in {"true", True}:
                return jsonify({"error": "Google email is not verified"}), 401

            email = str(token_info.get("email", "")).strip().lower()
            name = str(token_info.get("name") or token_info.get("given_name") or email).strip()
            if not email:
                return jsonify({"error": "No email returned by Google"}), 401

            user = User.query.filter_by(email=email).first()
            if user is None:
                user = User(email=email, display_name=name or email, role="lab")
                user.set_password(_sec.token_urlsafe(32))
                db.session.add(user)
                db.session.flush()
            token_str = AuthToken.generate(user.id)
            return jsonify({"token": token_str, "user": user.to_dict()})
        except Exception:
            return jsonify({"error": "Google token verification failed"}), 502

    @app.route("/api/auth/oauth/google")
    def oauth_google_start():
        config_error, client_id, _ = _google_client_config_error()
        if config_error:
            return jsonify({"error": config_error}), 503
        state = _sec.token_urlsafe(16)
        _oauth_states[state] = "google"
        redirect_uri = os.environ.get(
            "GOOGLE_REDIRECT_URI",
            "http://localhost:5003/api/auth/oauth/google/callback",
        )
        params = urllib.parse.urlencode({
            "client_id": client_id,
            "redirect_uri": redirect_uri,
            "response_type": "code",
            "scope": "openid email profile",
            "state": state,
            "access_type": "online",
        })
        return redirect(f"https://accounts.google.com/o/oauth2/v2/auth?{params}")

    @app.route("/api/auth/oauth/google/callback")
    def oauth_google_callback():
        code = request.args.get("code", "")
        state = request.args.get("state", "")
        frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
        if not code or state not in _oauth_states:
            return redirect(f"{frontend_url}?oauth_error=invalid_state")
        _oauth_states.pop(state)
        redirect_uri = os.environ.get(
            "GOOGLE_REDIRECT_URI",
            "http://localhost:5003/api/auth/oauth/google/callback",
        )
        try:
            token_res = _http.post(
                "https://oauth2.googleapis.com/token",
                data={
                    "code": code,
                    "client_id": os.environ["GOOGLE_CLIENT_ID"],
                    "client_secret": os.environ["GOOGLE_CLIENT_SECRET"],
                    "redirect_uri": redirect_uri,
                    "grant_type": "authorization_code",
                },
                timeout=10,
            )
            if not token_res.ok:
                return redirect(f"{frontend_url}?oauth_error=token_exchange_failed")
            access_token = token_res.json().get("access_token", "")
            info_res = _http.get(
                "https://www.googleapis.com/oauth2/v2/userinfo",
                headers={"Authorization": f"Bearer {access_token}"},
                timeout=10,
            )
            if not info_res.ok:
                return redirect(f"{frontend_url}?oauth_error=userinfo_failed")
            info = info_res.json()
            email = info.get("email", "").lower()
            name = info.get("name") or info.get("given_name") or email
        except Exception:
            return redirect(f"{frontend_url}?oauth_error=request_failed")
        return _oauth_finish(email, name, "google", frontend_url)

    # ── Group / lab routes ─────────────────────────────────────────────────────

    @app.route("/api/me/role", methods=["PATCH"])
    @login_required
    def update_my_role():
        user = _get_current_user()
        data = request.get_json(silent=True) or {}
        role = str(data.get("role", "")).strip().lower()
        if role not in ("manager", "lab"):
            return jsonify({"error": "Role must be 'manager' or 'lab'"}), 400
        user.role = role
        db.session.commit()
        return jsonify({"user": user.to_dict()})

    @app.route("/api/group", methods=["GET"])
    @login_required
    def get_group_info():
        user = _get_current_user()
        if user.group_id is None:
            return jsonify({"group": None})

        group = LabGroup.query.get(user.group_id)
        if group is None:
            return jsonify({"group": None})

        return jsonify({"group": group.to_dict()})

    @app.route("/api/group", methods=["POST"])
    @login_required
    def create_or_update_group_info():
        user = _get_current_user()
        if user.role != "manager":
            return jsonify({"error": "Only managers can update group details"}), 403

        data = request.get_json(silent=True) or {}
        lab_name = str(data.get("lab_name", "")).strip()
        group_name = str(data.get("group_name", "")).strip()

        if not lab_name or not group_name:
            return jsonify({"error": "Lab name and group name are required"}), 400

        group = LabGroup.query.get(user.group_id) if user.group_id else None
        if group is None:
            group = LabGroup(
                lab_name=lab_name,
                group_name=group_name,
                owner_user_id=user.id,
            )
            db.session.add(group)
            db.session.flush()
            user.group_id = group.id
        else:
            group.lab_name = lab_name
            group.group_name = group_name

        db.session.commit()
        return jsonify({"group": group.to_dict(), "user": user.to_dict()})

    @app.route("/api/group/members", methods=["GET"])
    @login_required
    def list_group_members():
        user = _get_current_user()
        if user.group_id is None:
            return jsonify({"members": []})

        members = User.query.filter_by(group_id=user.group_id).order_by(User.created_at.asc()).all()
        return jsonify({"members": [member.to_dict() for member in members]})

    @app.route("/api/group/invites", methods=["GET"])
    @login_required
    def list_group_invites():
        user = _get_current_user()
        if user.group_id is None:
            return jsonify({"invites": []})

        invites = LabInvite.query.filter_by(
            group_id=user.group_id,
            status="pending",
        ).order_by(LabInvite.created_at.desc(), LabInvite.id.desc()).all()

        return jsonify({"invites": [invite.to_dict() for invite in invites]})

    @app.route("/api/my-invites", methods=["GET"])
    @login_required
    def list_my_invites():
        user = _get_current_user()
        invites = LabInvite.query.filter_by(
            email=user.email,
            status="pending",
        ).order_by(LabInvite.created_at.desc(), LabInvite.id.desc()).all()

        payload = []
        for invite in invites:
            group = LabGroup.query.get(invite.group_id)
            payload.append({
                **invite.to_dict(),
                "group": group.to_dict() if group else None,
            })

        return jsonify({"invites": payload})

    @app.route("/api/my-invites/<int:invite_id>/accept", methods=["POST"])
    @login_required
    def accept_my_invite(invite_id):
        user = _get_current_user()
        invite = LabInvite.query.get_or_404(invite_id)

        if invite.email != user.email:
            return jsonify({"error": "That invite does not belong to you"}), 403
        if invite.status != "pending":
            return jsonify({"error": "That invite is no longer pending"}), 400
        if user.group_id is not None:
            return jsonify({"error": "Leave your current group before accepting another invite"}), 409

        group = LabGroup.query.get(invite.group_id)
        if group is None:
            return jsonify({"error": "That group no longer exists"}), 404

        user.group_id = invite.group_id
        if invite.role in ("manager", "lab"):
            user.role = invite.role
        invite.status = "accepted"
        invite.accepted_at = datetime.utcnow()
        db.session.commit()

        return jsonify({
            "invite": invite.to_dict(),
            "group": group.to_dict(),
            "user": user.to_dict(),
        })

    @app.route("/api/my-invites/<int:invite_id>/decline", methods=["POST"])
    @login_required
    def decline_my_invite(invite_id):
        user = _get_current_user()
        invite = LabInvite.query.get_or_404(invite_id)

        if invite.email != user.email:
            return jsonify({"error": "That invite does not belong to you"}), 403
        if invite.status != "pending":
            return jsonify({"error": "That invite is no longer pending"}), 400

        invite.status = "declined"
        db.session.commit()
        return jsonify({"invite": invite.to_dict()})

    @app.route("/api/group/members/<int:member_id>", methods=["DELETE"])
    @login_required
    def remove_group_member(member_id):
        user = _get_current_user()
        if user.role != "manager":
            return jsonify({"error": "Only managers can remove members"}), 403
        if user.group_id is None:
            return jsonify({"error": "You are not in a group"}), 400

        target = User.query.get_or_404(member_id)
        if target.group_id != user.group_id:
            return jsonify({"error": "That user is not in your group"}), 404
        if target.id == user.id:
            return jsonify({"error": "You cannot remove yourself"}), 400

        target.group_id = None
        db.session.commit()
        return jsonify({"removed_user_id": member_id})

    @app.route("/api/group/invite", methods=["POST"])
    @login_required
    def invite_group_member():
        user = _get_current_user()
        if user.role != "manager":
            return jsonify({"error": "Only managers can send invites"}), 403
        if user.group_id is None:
            return jsonify({"error": "Create your group first"}), 400

        data = request.get_json(silent=True) or {}
        email = str(data.get("email", "")).strip().lower()
        role = str(data.get("role", "lab")).strip().lower()

        if not email:
            return jsonify({"error": "Email is required"}), 400
        if role not in ("manager", "lab"):
            role = "lab"

        existing_user = User.query.filter_by(email=email).first()
        if existing_user and existing_user.group_id == user.group_id:
            return jsonify({"error": "That user is already in your group"}), 409
        if existing_user and existing_user.group_id is not None and existing_user.group_id != user.group_id:
            return jsonify({"error": "That user is already in another group"}), 409

        invite = LabInvite.query.filter_by(
            email=email,
            group_id=user.group_id,
            status="pending",
        ).first()

        if invite is None:
            invite = LabInvite(
                email=email,
                group_id=user.group_id,
                invited_by_user_id=user.id,
                role=role,
                status="pending",
            )
            db.session.add(invite)
        else:
            invite.role = role
            invite.invited_by_user_id = user.id

        db.session.commit()
        return jsonify({
            "invite": invite.to_dict(),
            "email_sent": False,
            "email_message": "Email delivery is disabled.",
        })

    # ── Purchase history ────────────────────────────────────────────────────────

    @app.route("/api/purchase-history", methods=["GET"])
    def list_purchase_history():
        entries = PurchaseHistory.query.order_by(
            PurchaseHistory.ordered_at.desc(), PurchaseHistory.id.desc()
        ).all()
        return jsonify([entry.to_dict() for entry in entries])

    @app.route("/api/purchase-history/<int:entry_id>", methods=["DELETE"])
    def delete_purchase_history_entry(entry_id):
        entry = PurchaseHistory.query.get_or_404(entry_id)
        db.session.delete(entry)
        db.session.commit()
        return jsonify({"success": True})

    @app.route("/api/purchase-history/<int:entry_id>/order-again", methods=["POST"])
    def order_again_from_history(entry_id):
        entry = PurchaseHistory.query.get_or_404(entry_id)
        reorder_quantity = max(1.0, float(entry.quantity_ordered or 0) or 0)

        item = None
        if entry.inventory_item_id is not None:
            item = InventoryItem.query.get(entry.inventory_item_id)

        if item is None:
            item = InventoryItem.query.filter_by(
                name=entry.item_name,
                catalog_number=entry.catalog_number,
            ).order_by(InventoryItem.updated_at.desc(), InventoryItem.id.desc()).first()

        if item is None:
            item = InventoryItem(
                name=entry.item_name,
                category=entry.category,
                catalog_number=entry.catalog_number,
                vendor=entry.vendor,
                requested_by=entry.requested_by,
                lab_name=entry.lab_name,
                actual_quantity=0,
                desired_quantity=reorder_quantity,
                order_quantity=reorder_quantity,
                need_to_order=True,
                order_status="needs",
                in_process_quantity=0,
            )
            db.session.add(item)
        else:
            item.vendor = item.vendor or entry.vendor
            item.requested_by = item.requested_by or entry.requested_by
            item.lab_name = item.lab_name or entry.lab_name
            item.category = item.category or entry.category
            item.desired_quantity = float(item.desired_quantity or 0) + reorder_quantity
            recalculate_need_fields(item, preserve_in_process=True)

        recalculate_need_fields(item, preserve_in_process=True)
        db.session.commit()
        return jsonify({"item": item.to_dict(), "quantity": reorder_quantity})

    @app.route("/api/purchase-history/import", methods=["POST"])
    def import_purchase_history_excel():
        file = request.files.get("file")
        if file is None or file.filename == "":
            return jsonify({"error": "No file uploaded"}), 400

        filename = file.filename.lower()
        if not (filename.endswith(".xlsx") or filename.endswith(".xlsm") or filename.endswith(".xltx")):
            return jsonify({"error": "Please upload an Excel .xlsx file"}), 400

        replace_existing = str(request.form.get("replace_existing", "false")).lower() == "true"

        workbook = load_workbook(file, data_only=True)

        if replace_existing:
            PurchaseHistory.query.delete()

        imported = 0
        skipped = 0
        skipped_missing_item_name = 0
        ignored_empty_rows = 0
        found_header_in_any_sheet = False

        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            header_index = None
            column_indexes = None
            for i, row in enumerate(rows):
                if row and any(cell is not None and str(cell).strip() != "" for cell in row):
                    maybe_indexes = find_column_indexes_for_aliases(row, PURCHASE_HISTORY_HEADER_ALIASES)
                    if "item_name" in maybe_indexes:
                        header_index = i
                        column_indexes = maybe_indexes
                        break

            if header_index is not None and column_indexes is not None:
                found_header_in_any_sheet = True

            data_start_index = (header_index + 1) if header_index is not None else 0

            for row in rows[data_start_index:]:
                if row is None:
                    ignored_empty_rows += 1
                    continue

                if all(cell is None or str(cell).strip() == "" for cell in row):
                    ignored_empty_rows += 1
                    continue

                if column_indexes is None:
                    if not is_likely_history_data_row(row):
                        ignored_empty_rows += 1
                        continue

                    # Fallback for headerless sheets:
                    # A: ordered date, B: requested by, C: lab/group, E: vendor,
                    # F: item description, G: status, H: catalog/order reference.
                    raw_item_name = row[5] if len(row) > 5 else ""
                    catalog_number = clean_text(row[7] if len(row) > 7 else None)
                    vendor = clean_text(row[4] if len(row) > 4 else None)
                    requested_by = clean_text(row[1] if len(row) > 1 else None)
                    lab_name = clean_text(row[2] if len(row) > 2 else None)
                    status_text = clean_text(row[6] if len(row) > 6 else None) or "ordered"
                    ordered_at_value = parse_datetime_cell(row[0] if len(row) > 0 else None)
                    received_at_value = None
                    quantity_ordered = 0
                    quantity_received = 0
                else:
                    raw_item_name = get_row_value(row, column_indexes, "item_name", "")
                    catalog_number = clean_text(get_row_value(row, column_indexes, "catalog_number"))
                    vendor = clean_text(get_row_value(row, column_indexes, "vendor"))
                    requested_by = clean_text(get_row_value(row, column_indexes, "requested_by"))
                    lab_name = clean_text(get_row_value(row, column_indexes, "lab_name"))
                    status_text = clean_text(get_row_value(row, column_indexes, "status")) or "ordered"
                    ordered_at_value = parse_datetime_cell(get_row_value(row, column_indexes, "ordered_at"))
                    if ordered_at_value is None and "ordered_at" not in column_indexes:
                        # Some sheets include headers for item columns but leave the date column unlabeled.
                        # In those cases, use column A as ordered date when parseable.
                        ordered_at_value = parse_datetime_cell(row[0] if len(row) > 0 else None)
                    received_at_value = parse_datetime_cell(get_row_value(row, column_indexes, "received_at"))
                    quantity_ordered = to_float(get_row_value(row, column_indexes, "quantity_ordered", 0), 0)
                    quantity_received = to_float(get_row_value(row, column_indexes, "quantity_received", 0), 0)

                item_name = str(raw_item_name).strip() if raw_item_name is not None else ""
                if item_name == "":
                    skipped += 1
                    skipped_missing_item_name += 1
                    continue

                entry = PurchaseHistory(
                    item_name=item_name,
                    category=clean_text(get_row_value(row, column_indexes, "category")) if column_indexes else None,
                    catalog_number=catalog_number,
                    vendor=vendor,
                    requested_by=requested_by,
                    lab_name=lab_name,
                    quantity_ordered=quantity_ordered,
                    quantity_received=quantity_received,
                    status=status_text,
                    ordered_at=ordered_at_value,
                    received_at=received_at_value,
                )
                db.session.add(entry)
                imported += 1

        if imported == 0 and not found_header_in_any_sheet:
            return jsonify({"error": "Could not detect history rows. Expected a header with Item/Name or a table where column A is date and column F is item description."}), 400

        db.session.commit()

        entries = PurchaseHistory.query.order_by(
            PurchaseHistory.ordered_at.desc(), PurchaseHistory.id.desc()
        ).all()

        return jsonify(
            {
                "imported": imported,
                "skipped": skipped,
                "skipped_missing_item_name": skipped_missing_item_name,
                "ignored_empty_rows": ignored_empty_rows,
                "total": len(entries),
                "entries": [entry.to_dict() for entry in entries],
            }
        )

    @app.route("/api/inventory", methods=["GET"])
    def list_inventory():
        items = InventoryItem.query.order_by(InventoryItem.name).all()
        return jsonify([item.to_dict() for item in items])

    @app.route("/api/inventory", methods=["POST"])
    def create_inventory_item():
        data = request.get_json() or {}
        item = InventoryItem(
            name=str(data.get("name", "")).strip(),
            category=clean_text(data.get("category")),
            location_stored=clean_text(data.get("location_stored")),
            catalog_number=clean_text(data.get("catalog_number")),
            vendor=clean_text(data.get("vendor")),
            requested_by=clean_text(data.get("requested_by")),
            lab_name=clean_text(data.get("lab_name")),
            actual_quantity=data.get("actual_quantity", 0),
            desired_quantity=data.get("desired_quantity"),
            need_to_order=False,
            order_quantity=0,
            order_status="in_stock",
            in_process_quantity=0,
        )
        recalculate_need_fields(item)
        db.session.add(item)
        db.session.commit()
        return jsonify(item.to_dict()), 201

    @app.route("/api/inventory/import", methods=["POST"])
    def import_inventory_excel():
        file = request.files.get("file")
        if file is None or file.filename == "":
            return jsonify({"error": "No file uploaded"}), 400

        filename = file.filename.lower()
        if not (filename.endswith(".xlsx") or filename.endswith(".xlsm") or filename.endswith(".xltx")):
            return jsonify({"error": "Please upload an Excel .xlsx file"}), 400

        replace_existing = str(request.form.get("replace_existing", "true")).lower() == "true"

        workbook = load_workbook(file, data_only=True)

        if replace_existing:
            InventoryItem.query.delete()
            existing_keys = set()
        else:
            existing_keys = {
                make_item_key(item.name, item.catalog_number, item.location_stored)
                for item in InventoryItem.query.all()
            }

        imported = 0
        skipped = 0
        seen_file_keys = set()

        found_header_in_any_sheet = False

        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            header_index = None
            column_indexes = None
            for i, row in enumerate(rows):
                if row and any(cell is not None and str(cell).strip() != "" for cell in row):
                    maybe_indexes = find_column_indexes(row)
                    if "name" in maybe_indexes:
                        header_index = i
                        column_indexes = maybe_indexes
                        break

            if header_index is None or column_indexes is None:
                continue

            found_header_in_any_sheet = True

            for row in rows[header_index + 1:]:
                if row is None:
                    continue

                raw_name = get_row_value(row, column_indexes, "name", "")
                name = str(raw_name).strip() if raw_name is not None else ""
                if name == "":
                    skipped += 1
                    continue

                catalog_number = str(get_row_value(row, column_indexes, "catalog_number", "")).strip() or None
                location_stored = str(get_row_value(row, column_indexes, "location_stored", "")).strip() or None
                item_key = make_item_key(name, catalog_number, location_stored)
                if item_key in seen_file_keys or item_key in existing_keys:
                    skipped += 1
                    continue
                seen_file_keys.add(item_key)
                existing_keys.add(item_key)

                raw_category = get_row_value(row, column_indexes, "category", "")
                category = str(raw_category).strip() if raw_category else None

                item = InventoryItem(
                    name=name,
                    category=category,
                    catalog_number=catalog_number,
                    location_stored=location_stored,
                    actual_quantity=to_float(get_row_value(row, column_indexes, "actual_quantity", 0), 0),
                    desired_quantity=to_float(get_row_value(row, column_indexes, "desired_quantity", 0), 0),
                    need_to_order=False,
                    order_quantity=0,
                    order_status="in_stock",
                    in_process_quantity=0,
                )
                recalculate_need_fields(item)
                db.session.add(item)
                imported += 1

        if not found_header_in_any_sheet:
            return jsonify({"error": "Could not detect header row with an Item/Name column in any sheet"}), 400

        db.session.commit()

        items = InventoryItem.query.order_by(InventoryItem.name).all()
        return jsonify(
            {
                "imported": imported,
                "skipped": skipped,
                "total": len(items),
                "items": [item.to_dict() for item in items],
            }
        )

    @app.route("/api/inventory/<int:item_id>", methods=["PUT"])
    def update_inventory_item(item_id):
        item = InventoryItem.query.get_or_404(item_id)
        data = request.get_json() or {}
        item.name = str(data.get("name", item.name)).strip()
        item.category = clean_text(data.get("category"))
        item.location_stored = clean_text(data.get("location_stored"))
        item.catalog_number = clean_text(data.get("catalog_number"))
        item.vendor = clean_text(data.get("vendor"))
        item.requested_by = clean_text(data.get("requested_by"))
        item.lab_name = clean_text(data.get("lab_name"))
        item.actual_quantity = data.get("actual_quantity", item.actual_quantity)
        item.desired_quantity = data.get("desired_quantity", item.desired_quantity)
        recalculate_need_fields(item, preserve_in_process=True)
        db.session.commit()
        return jsonify(item.to_dict())

    @app.route("/api/inventory/<int:item_id>", methods=["DELETE"])
    def delete_inventory_item(item_id):
        item = InventoryItem.query.get_or_404(item_id)
        db.session.delete(item)
        db.session.commit()
        return jsonify({"success": True})

    @app.route("/api/inventory/<int:item_id>/take", methods=["POST"])
    def take_inventory_item(item_id):
        item = InventoryItem.query.get_or_404(item_id)
        data = request.get_json() or {}
        amount = float(data.get("amount", 1))
        if amount <= 0:
            return jsonify({"error": "Amount must be greater than 0"}), 400

        item.actual_quantity = max(0, float(item.actual_quantity or 0) - amount)
        recalculate_need_fields(item, preserve_in_process=True)

        db.session.commit()
        return jsonify(item.to_dict())

    @app.route("/api/inventory/<int:item_id>/request-more", methods=["POST"])
    def request_more_inventory_item(item_id):
        item = InventoryItem.query.get_or_404(item_id)
        data = request.get_json() or {}
        requested_amount = float(data.get("requested_amount", 1))
        if requested_amount <= 0:
            return jsonify({"error": "Requested amount must be greater than 0"}), 400

        # Interpret "request more" as increasing target stock level.
        item.desired_quantity = float(item.desired_quantity or 0) + requested_amount
        recalculate_need_fields(item, preserve_in_process=True)
        db.session.commit()
        return jsonify(item.to_dict())

    @app.route("/api/inventory/<int:item_id>/mark-ordered", methods=["POST"])
    def mark_item_ordered(item_id):
        item = InventoryItem.query.get_or_404(item_id)
        need = max(0.0, float(item.desired_quantity or 0) - float(item.actual_quantity or 0))
        if need <= 0:
            return jsonify({"error": "Item does not currently need ordering"}), 400

        item.in_process_quantity = need
        item.order_status = "in_process"
        item.ordered_at = datetime.utcnow()
        item.last_order_date = item.ordered_at
        item.need_to_order = False
        create_purchase_history_entry(item, need, item.ordered_at)
        db.session.commit()
        return jsonify(item.to_dict())

    @app.route("/api/inventory/<int:item_id>/mark-received", methods=["POST"])
    def mark_item_received(item_id):
        item = InventoryItem.query.get_or_404(item_id)
        receive_amount = float(item.in_process_quantity or 0)
        if receive_amount <= 0:
            return jsonify({"error": "No in-process order to receive"}), 400

        item.actual_quantity = float(item.actual_quantity or 0) + receive_amount
        item.in_process_quantity = 0
        item.received_at = datetime.utcnow()
        recalculate_need_fields(item)

        history_entry = find_open_purchase_history_entry(item.id)
        if history_entry is None:
            history_entry = create_purchase_history_entry(item, receive_amount, item.received_at)
        history_entry.quantity_received = receive_amount
        history_entry.received_at = item.received_at
        history_entry.status = "received"
        db.session.commit()
        return jsonify(item.to_dict())

    @app.route("/api/inventory/request-need", methods=["POST"])
    def request_need():
        data = request.get_json(silent=True) or {}
        name = clean_text(data.get("name"))
        if not name:
            return jsonify({"error": "Item name is required"}), 400
        qty = max(0.0, to_float(data.get("quantity"), 0.0))
        item = InventoryItem(
            name=name,
            catalog_number=clean_text(data.get("catalog_number")),
            vendor=clean_text(data.get("vendor")),
            requested_by=clean_text(data.get("requested_by")),
            lab_name=clean_text(data.get("lab_name")),
            actual_quantity=0,
            desired_quantity=qty,
            order_quantity=qty,
            need_to_order=True,
            order_status="needs",
        )
        db.session.add(item)
        db.session.commit()
        return jsonify(item.to_dict()), 201
